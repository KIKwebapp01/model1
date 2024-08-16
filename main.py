import streamlit as st
import pandas as pd
import os
from opt import execute_optimization
from datetime import datetime, time, timedelta
from openpyxl import load_workbook


df = None           # 読み込んだファイルのDataFrame
is_solved = False   # スケジュールを立案したか
input_file = None


def read_data():
    global df, is_solved, input_file

    st.markdown(
        """
        #### スケジュールで使用するデータを読み込みます
        """
    )
    input_mode = st.radio(
        "読込方法：",
        ('データをアップロードする', 'サンプルデータを読み込む'), horizontal=True
    )
    if input_mode == 'データをアップロードする':
        input_file = st.file_uploader("Choose a file")
        if input_file is not None:
            df = pd.read_excel(input_file, index_col='ID')
            df['納期'] = df['納期'].dt.date
            # 読み込んだDataFrameを表示
            st.write("読み込んだデータ:")
            st.dataframe(df)
            is_solved = False
    else:
        st.write("読み込むサンプルファイルを選択")

        # カレントフォルダ内のsampledataフォルダからExcelファイル一覧を取得
        sampledata_folder = "sampledata"
        file_list = [f for f in os.listdir(sampledata_folder) if f.lower().endswith(".xlsx")]

        # ファイル選択用の選択ボックスを表示
        input_file = st.selectbox("Excelファイルを選択してください", file_list)

        # 選択されたExcelファイルをDataFrameとして読み込み
        if input_file:
            file_path = os.path.join(sampledata_folder, input_file)
            df = pd.read_excel(file_path, index_col='ID')
            df['納期'] = df['納期'].dt.date

            # 読み込んだDataFrameを表示
            st.write("読み込んだデータ:")
            st.dataframe(df)
            is_solved = False


def make_schedule():
    global is_solved, input_file

    if df is None:
        st.write("まず，データを読み込んで下さい．")
        return
    with st.expander(f"読込データ: {input_file}（折り畳みを解除して確認できます）"):
        st.dataframe(df)

    # スケジュール作成実行
    if st.button("スケジュール作成実行"):
        ret_df = execute_optimization(df)
        if ret_df is None:
            st.write('最適なスケジュールが見つかりませんでした．入力ファイルを確認してください．')
        else:
            is_solved = True

    if is_solved:  # スケジュール立案済みのとき
        st.dataframe(ret_df)


def change_settings():
    st.markdown(
        """
        #### スケジュールで使用する設定を変更できます
        """
    )
    # settings.xlsxから設定を読み込む
    wb = load_workbook("settings.xlsx")
    ws = wb.active

    # セルB2から読み込んだ時刻を基準時刻 sとする．datetime型
    now = datetime.now()
    today = now.date()
    s = datetime.strptime(f"{now.year}-{now.month}-{now.day}-{ws.cell(row=2, column=2).value}", '%Y-%m-%d-%H:%M')

    # A行目から読み込んだ各時刻を初期値として，選択ボックスを表示
    t = [now.time()] * 6       # 各要素：datetime.time型
    for i in range(6):
        minute_to_add = ws.cell(row=1, column=i + 2).value
        t[i] = (s + timedelta(minutes=float(minute_to_add))).time()
    # AM開始時刻
    t[0] = st.time_input('AM開始時刻', t[0], step=60)
    s = datetime.combine(today, t[0])
    ws.cell(row=2, column=2).value = datetime.combine(today, t[0]).strftime('%H:%M')
    # それ以外の時刻
    for i, label in enumerate(['AM終了時刻', 'PM1開始時刻', 'PM1終了時刻', 'PM2開始時刻', 'PM2終了時刻'], start=1):
        t[i] = st.time_input(label, t[i], step=60)
        ws.cell(row=1, column=i + 2).value = (datetime.combine(today, t[i]) - s).seconds // 60

    # 設定値を setting.xlsxに保存する
    wb.save('settings.xlsx')


def main():
    # 画面全体の設定
    st.set_page_config(
        page_title="スケジュール自動作成アプリ",
        page_icon="🖥️",
        layout="centered",
    )

    tab1, tab2, tab3 = st.tabs(["データ読込", "スケジュール作成", "設定変更"])
    with tab1:
        read_data()
    with tab2:
        make_schedule()
    with tab3:
        change_settings()


if __name__ == "__main__":
    main()
